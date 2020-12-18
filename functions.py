import numpy as np
# import matplotlib.pyplot as plt
import openpyxl as px
from itertools import groupby


# Funcion para lectura y obtencion de datos de inicio
# La funcion devuelve:
# demanda_clientes: matriz de demandas de los clientes por periodo y por producto
# capacidad_vehiculos_p: matriz de capacidades de los vehiculos de primer escalon para cada producto
# capacidad_vehiculos_s: matriz de capacidades de los vehiculos de segundo escalon para cada producto
# capacidad_cr: matriz de capacidades de los centros regionales en cada periodo
# capacidad_cl: matriz de capacidades de los centros locales en cada periodo
# inventario: matriz de inventarios iniciales de cada producto en cada centro regional
# costo_inventario: matriz de costos de mantener en inventario cada producto en cada centro regional
# costo_instalaciones_cr: matriz de costos de habilitacion de los centros regionales
# costo_instalaciones_cl: matriz de costos de habilitacion de los centros locales
# costo_vehiculos_p: matriz de costos de utilizacion de los vehiculos de primer escalon
# costo_vehiculos_s: matriz de costos de utilizacion de los vehiculos de segundo escalon
# costo_compraproductos: matriz de costos de la compra de cada producto en cada periodo
# costo_transporte: matriz de costos unitarios de transporte de cada producto en cada periodo para cada centro regional
# costo_rutas_p: matriz de costos de rutas entre centros regionales y centros locales
# costo_rutas_s: matriz de costos de rutas entre centrol locales y clientes
# costo_humano: matrices de costo de sufrimiento humano, cada matriz corresponde a un periodo.
def read_data(n_clientes, n_productos, n_periodos, n_vehiculos_p, n_vehiculos_s, n_centrosregionales, n_centroslocales):
    # total_columnas = variable que me permite moverme entre los indices de las hojas de datos
    total_columnas = n_productos*n_periodos

    # lectura y obtencion de datos de los clientes
    datos = px.load_workbook('datos4.xlsx')                                          # carga de la hoja de excel de datos
    hoja_clientes = datos['clientes']                                               # seleccionar la hoja clientes como hoja activa
    # Obtencion de las demandas de la tabla de la hoja clientes segun la cantidad de clientes, productos y periodos
    demanda_clientes = [[hoja_clientes.cell(row=i, column=j).value for j in range(2, 2+total_columnas)] for i in range(3, 3+n_clientes)]

    # lectura y obtencion de datos de los vehiculos de primer y segundo nivel
    hoja_vehiculos = datos['vehiculos']                                             # seleccionar la hoja vehiculos como hoja activa
    # Obtencion de las demandas de los vehiculos de primer nivel de la primera tabla en la hoja vehiculos segun la cantidad de vehiculos de primer nivel y periodos
    capacidad_vehiculos_p = [[hoja_vehiculos.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(2, 2+n_vehiculos_p)]
    # Obtencion de las demandas de los vehiculos de segundo nivel de la segunda tabla en la hoja vehiculos segun la cantidad de periodos, vehiculos de primer y segundo nivel
    capacidad_vehiculos_s = [[hoja_vehiculos.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(3+n_vehiculos_p, 3+n_vehiculos_p+n_vehiculos_s)]

    # lectura y obtencion de datos de las instalaciones de primer y segundo nivel
    hoja_instalaciones = datos['instalaciones']                                     # seleccionar la hoja instalaciones como hoja activa
    # Obtencion de las capacidades de los centros de primer nivel de la primera tabla en la hoja instalaciones segun la cantidad de centros regionales y periodos
    capacidad_cr = [[hoja_instalaciones.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(3, 3+n_centrosregionales)]
    # Obtencion de las capacidades de los centros de segundo nivel de la segunda tabla en la hoja instalaciones segun la cantidad de centros locales, centros regionales y periodos
    capacidad_cl = [[hoja_instalaciones.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(5+n_centrosregionales, 5+n_centrosregionales+n_centroslocales)]

    # lectura y obtencion de datos del inventario inicial para el primer escalon
    hoja_inventario = datos['inventario']                                           # seleccionar la hoja inventario como hoja activa
    # Obtencion de los datos en la hoja seleccionada segun a cantidad de centros regionales y productos
    inventario = [[hoja_inventario.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(2, 2+n_centrosregionales)]

    # lectura y obtencion de datos de costo de mantener el inventario
    hoja_costo_inventario = datos['costo_inventario']                               # seleccionar la hoja costo_inventario como hoja activa
    # Obtencion de los costos en la hoja seleccionada segun la cantidad de centros regionales y productos
    costo_inventario = [[hoja_costo_inventario.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(2, 2+n_centrosregionales)]

    # lectura y obtencion de datos de costo de habilitar las instalaciones de primer y segundo nivel
    hoja_costo_instalaciones = datos['costo_instalaciones']                         # seleccionar la hoja costo_instalaciones como hoja activa
    # obtencion de los costos de los centros regionales en la hoja seleccionada
    costo_instalaciones_cr = [hoja_costo_instalaciones.cell(row=i, column=2).value for i in range(3, 3+n_centrosregionales)]
    # obtencion de los costos de los centros locales en la hoja seleccionada
    costo_instalaciones_cl = [hoja_costo_instalaciones.cell(row=i, column=2).value for i in range(5+n_centrosregionales, 5+n_centrosregionales+n_centroslocales)]

    # lectura y obtencion de datos de costo de uso de vehiculos de primer y segundo nivel
    hoja_costo_vehiculos = datos['costo_vehiculos']                                 # seleccionar la hoja costo_vehiculos como hoja activa
    # Obtencion de los costos de uso de los vehiculos de primer nivel en la hoja seleccionada
    costo_vehiculos_p = [hoja_costo_vehiculos.cell(row=i, column=2).value for i in range(2, 2+n_vehiculos_p)]
    # Obtencion de los costos de uso de los vehiculos de segundo nivel en la hoja seleccionada
    costo_vehiculos_s = [hoja_costo_vehiculos.cell(row=i, column=2).value for i in range(3+n_vehiculos_p, 3+n_vehiculos_p+n_vehiculos_s)]

    # lectura y obtencion de datos de costo de comprar productos en cada periodo
    hoja_costo_compraproductos = datos['costo_compraproductos']                     # seleccionar la hoja costo_compraproductos como hoja activa
    # Obtencion de los costos de compra de productos segun la cantidad de productos y el numero de periodos
    costo_compraproductos = [[hoja_costo_compraproductos.cell(row=i, column=j).value for j in range(2, 2+n_periodos)] for i in range(2, 2+n_productos)]

    # lectura y obtencion de valores de costo unitario de transporte los centros regionales
    hoja_costo_transporte = datos['costo_transporte']                               # seleccionar la hoja costo_transporte como hoja activa
    # Obtencion de los costos de transporte de los centros regionales segun cantidad de productos y periodos
    costo_transporte = [[hoja_costo_transporte.cell(row=i, column=j).value for j in range(2, 2+total_columnas)] for i in range(3, 3+n_centrosregionales)]

    # lectura y obtencion de los valores de costo de las rutas de primer nivel
    hoja_costo_rutas_p = datos['costo_rutas_p']                                     # seleccionar la hoja costo_rutas_p como hoja activa
    # Obtencion de los costos de ruteo entre centros regionales y centros locales
    costo_rutas_p = [[hoja_costo_rutas_p.cell(row=i, column=j).value for j in range(2, 2+n_centrosregionales+n_centroslocales)] for i in range(2, 2+n_centrosregionales+n_centroslocales)]

    # lectura y obtencion de los valores de costo de las rutas de segundo nivel
    hoja_costo_rutas_s = datos['costo_rutas_s']                                     # seleccionar la hoja costo_rutas_s como hoja activa
    # Obtencion de los costos de ruteo entre centros locales y clientes
    costo_rutas_s = [[hoja_costo_rutas_s.cell(row=i, column=j).value for j in range(2, 2+n_centroslocales+n_clientes)] for i in range(2, 2+n_centroslocales+n_clientes)]

    # lectura y obtencion de los valores de costo de sufrimiento humano
    hoja_costo_humano = datos['costo_humanitario']                                  # seleccionar la hoja costo_humanitario como hoja activa
    costo_humano = np.zeros([n_periodos, n_clientes+n_centroslocales, n_clientes+n_centroslocales])  # matriz de 3 dimensiones llena con ceros que se rellenara con cada matriz que se leera de la hoja activa
    i0 = 3                                                                          # indice inicial para las filas
    j0 = i0-1                                                                       # indice inicial para las columnas
    for k in range(n_periodos):
        # Obtencion de los costos de sufrimiento humano por cada periodo entre centros locales y clientes
        costo_humano[k] = [[hoja_costo_humano.cell(row=i, column=j).value for j in range(j0, j0+n_centroslocales+n_clientes)] for i in range(i0, i0+n_centroslocales+n_clientes)]
        i0 = i0 + n_centroslocales + n_clientes + 2                                 # Actualizacion del indice de las filas

    return np.array(demanda_clientes), np.array(capacidad_vehiculos_p), np.array(capacidad_vehiculos_s), np.array(capacidad_cr), np.array(capacidad_cl), np.array(inventario), np.array(costo_inventario), np.array(costo_instalaciones_cr), np.array(costo_instalaciones_cl), np.array(costo_vehiculos_p), np.array(costo_vehiculos_s), np.array(costo_compraproductos), np.array(costo_transporte), np.array(costo_rutas_p), np.array(costo_rutas_s), costo_humano


# Funcion que binariza los valores que se pasen como parametro
# con el objetivo de llevar un control de la operacion de resta realizada en las capacidades de los centros y los vehiculos
# la funcion devuelve un nuevo valor ya sea 1 o 0 si el valor es positivo o negativo respectivamente
def binarize(val):
    if val > 0:
        return 1
    else:
        return 0


# Funcion que traduce en un diccionario la matriz que se pasa como parametro
# con el objetivo de llevar un registro temporal de las asignaciones
def dictionarize(mat):
    dicc = {}                                                               # inicializacion de un diccionario vacio
    for idx, val in enumerate(mat[1, :]):                                   # recorrer cada uno de los valores de la matriz y enumerarlos para llevar un control de indices
        if val not in dicc.keys():
            dicc[val] = [mat[0, idx]]                                       # si el valor no esta en el diccionario se guarda y se le asigna el contenido correspondiente de la matriz
        else:
            dicc[val].append(mat[0, idx])                                   # en caso de que el valor si este en el diccionario se adjunta el contenido correspondiente de la matriz al ya asignado anteriormente
    return dicc


def dictionarize_cr(mat):
    dicti = {}
    for cr in mat:
        if cr[0] not in dicti.keys():
            dicti[cr[0]] = cr[1]
    return dicti


# Funcion que permite traducir o mapear las asignaciones de la matriz que se pasa como parametro
# con el objetivo de poder almacenar y utilizar los centros habilitados en cada asignacion
# la funcion retorna
# n_centros_habs: numero de centros habilitados
# demanda_cl_np: arreglo con la demanda de los centros habilitados
# centros_habs: arreglo con los centros habilitados
def maping(demanda):
    n_centros_habs = len(demanda)
    demanda_cl_np = np.array([x[1] for x in demanda])
    centros_habs = [int(x[0]) for x in demanda]
    return n_centros_habs, demanda_cl_np, centros_habs


# Funcion para las asignaciones (Decision de localizacion-asignacion) recibe como parametros:
# n_asignar: numero de clientes o centros locales que seran asignados segun el nivel
# n_centros: numero de centros locales o centros regionales a los que se asigna segun el nivel
# periodo: periodo en el que se encuentre la asignacion
# n_productos: numero de productos
# capacidad_centro: capacidad del centro al cual se le esten asignando clientes u otros centros segun el nivel
# demanda: demanda del cliente o centro segun el nivel
# mapeo: vector con los centros habilitados, para el segundo escalon en el periodo 1 la longitud del mapeo es igual a 0
# La funcion devuelve la matriz de localizacion-asignacion y la demanda del centro local o regional segun el nivel y el periodo que se este trabajando

def asignaciones(n_asignar, n_centros, periodo, n_productos, capacidad_centro, demanda, mapeo, escalon):

    por_asignar = np.array(range(1, n_asignar + 1))                             # creacion de una lista con los clientes o centros que se asignaran
    centros = np.array(range(1, n_centros + 1))                                 # creacion de una lista con los centros a los que se asginaran
    asignacion_lv = np.array([por_asignar, np.zeros(len(por_asignar))])         # inicializacion de la matriz localizacion-asignacion
    intentos = 0                                                                # variable para el control de intentos de asignaciones
    rango = (periodo-1)*n_productos                                             # rango de indices para moverse a traves de la matriz de demanda
    copia_capacidad = np.copy(capacidad_centro)                                 # copia de la capacidad del centro para evitar modificaciones en las capacidades originales
    if (len(mapeo) > 0 and escalon == 2) or (len(mapeo) > 0 and escalon == 1):
        centro_temp = int(np.random.choice(mapeo))                              # selecciona un centro habilitado previamente en otro periodo
    else:
        centro_temp = int(np.random.choice(centros))                            # seleccion aleatoria del primer centro - habilitacion del primer centro
    # idx_c = 0
    while len(por_asignar) > 0:                                                 # mientras existan clientes o centros por asignar
        if intentos < 3:                                                        # si los intentos de asignacion son menores a 3
            asig_temp = np.random.choice(por_asignar)                           # selecciona un cliente o centro aleatorio para asignar
            if periodo > 1 and escalon == 1:                                    # si el periodo actual es diferente al inicial y el escalon a asignar es el primero
                resta_capacidad = copia_capacidad[centro_temp - 1, :] - demanda[asig_temp - 1, 0:n_productos]  # resta la capacidad del centro con la demanda del centro
                copia_capacidad[centro_temp - 1, :] = copia_capacidad[centro_temp - 1, :] - demanda[asig_temp - 1, 0:n_productos]  # actualiza la nueva capacidad del centro
            else:
                resta_capacidad = copia_capacidad[centro_temp-1, :]-demanda[asig_temp-1, rango:rango+n_productos]  # resta la capacidad del centro con la demanda del centro
                copia_capacidad[centro_temp-1, :] = copia_capacidad[centro_temp-1, :]-demanda[asig_temp-1, rango:rango+n_productos]   # actualiza la nueva capacidad del centro
            binvec = np.array([binarize(x) for x in resta_capacidad])           # se binariza la resta de la capacidad
            if binvec.all():                                                    # si la resta en la capacidad para todos los productos da un valor positivo
                idx_c = int(np.where(asignacion_lv[0, :] == asig_temp)[0])      # almacena el indice del centro
                asignacion_lv[1, idx_c] = centro_temp                           # guarda en la matriz de localizacion-asignacion el centro en la posicion del cliente o centro seleccionado
                idx_d = int(np.where(por_asignar == asig_temp)[0])              # almacena el indice del cliente o centro que ya fue asignado
                por_asignar = np.delete(por_asignar, [idx_d])                   # elimina de la lista el cliente o centro asignado
            else:                                                               # en caso de que la resta de la capacidad resulte negativa para al menos 1 valor
                intentos += 1                                                   # aumenta en 1 el numero de intentos
                if periodo > 1 and escalon == 1:                                # si el periodo actual es diferente al inicial y el escalon a asignar es el primero
                    copia_capacidad[centro_temp - 1, :] = copia_capacidad[centro_temp - 1, :] + demanda[asig_temp - 1, 0:n_productos]  # reestablece la capacidad del centro al momento antes de la resta
                else:
                    copia_capacidad[centro_temp - 1, :] = copia_capacidad[centro_temp - 1, :] + demanda[asig_temp - 1, rango:rango + n_productos]  # reestablece la capacidad del centro al momento antes de la resta
        else:                                                                   # al llegar al numero maximo de intentos
            if (len(mapeo) > 1 and escalon == 2) or (len(mapeo) > 1 and escalon == 1):
                idx_cl = np.where(mapeo == centro_temp)                         # almacenamos el indice del centro previamente habilitado en la lista de mapeados
                mapeo = np.delete(mapeo, [idx_cl])                              # lo eliminamos de la lista de centros mapeados
                idx_m = np.where(centros == centro_temp)                        # almacenamos el indice del centro previamente habilitado en la lista de centros
                centros = np.delete(centros, [idx_m])                           # lo eliminamos de la lista de centros
                centro_temp = int(np.random.choice(mapeo))                           # seleccionamos un nuevo centro de forma aleatoria
                intentos = 0                                                    # reestablecemos el numero de intentos
            elif (len(mapeo) == 1 and escalon == 2) or (len(mapeo) == 1 and escalon == 1):                              # si se usaron todos los centros y aun hay clientes por asignar
                idx_cl = np.where(mapeo == centro_temp)                         # seleccionamos el indice del centro que ya agoto su capacidad
                mapeo = np.delete(mapeo, [idx_cl])                              # eliminamos el centro de la lista de centros mapeados
                centro_temp = int(np.random.choice(centros))                         # seleccionamos o habilitamos un nuevo centro que no se haya usado
                intentos = 0                                                    # reiniciamos el numero de intentos
            else:
                idx_cl = np.where(centros == centro_temp)                       # seleccionamos el indice del centro que ya agoto su capacidad
                centros = np.delete(centros, [idx_cl])                          # eliminamos de la lista de centros el centro que ya fue agotado
                centro_temp = int(np.random.choice(centros))                         # seleccionamos o habilitamos un nuevo centro
                intentos = 0                                                    # reiniciamos el numero de intentos

    dicc = dictionarize(asignacion_lv)                                          # generamos un diccionario con las asignaciones realizadas donde la llave es el centro y los valores con los centros o clientes asignados
    demandaf = []                                                               # inicializamos un vector donde se almacenaran la demandas finales de los centros
    for centro, asignados in dicc.items():                                      # para cada centro y valores asignados a ese centro
        demandacentro = [centro]                                                # creamos un vector con el centro seleccionado
        suma = 0                                                                # inicializamos la suma de las demandas
        for asig in asignados:
            if periodo > 1 and escalon == 1:
                suma += demanda[int(asig) - 1, 0:n_productos]
            else:
                suma += demanda[int(asig) - 1, rango:rango + n_productos]       # sumamos las demandas de cada cliente o centro que fue asignado a ese centro
        demandacentro.append(suma)                                              # adjuntamos la demanda al vector que contiene las demandas del centro
        demandaf.append(demandacentro)                                          # adjuntamos las demandas del centro al vector que contiene las demandas de todos los centros

    return asignacion_lv, demandaf


# Funcion para el plan de rutas, recibe como parametros:
# asignacion_lv: matriz de asignacion-localizacion
# n_vehiculos: numero de vehiculos
# demanda: demanda de clientes o centros segun corresponda el nivel
# periodo: periodo en el que se encuentre la asignacion de rutas
# n_productos: numero de productos
# La funcion devuelve una lista de listas con el plan de rutas del periodo que se este trabajando

def rutas(asignacion_lv, n_vehiculos, capacidad_vehiculos, demanda, periodo, n_productos, escalon):
    rutas_lv = []                                                         # inicializacion del vector de vetores de rutas
    vehiculos = list(range(1, n_vehiculos+1))                             # creacion de una lista de vehiculos con los vehiculos existentes
    dicci_asignacion = dictionarize(asignacion_lv)                        # generacion de un diccionario con la matriz de asignacion-localizacion
    rango = (periodo-1)*n_productos                                       # rango de indices para moverse a travesde la matriz de demandas
    capacidad_vehiculos_copy = np.copy(capacidad_vehiculos)
    for centro, asignados in dicci_asignacion.items():                    # por cada centro y sus respectivas asignaciones
        idx_c = 0                                                         # inicializacion del indice para recorrer los centros o clientes asignados
        vehiculo_temp = np.random.choice(vehiculos)                       # seleccionamos un vehiculo aleatorio
        ruta_temp = [int(centro), vehiculo_temp, 0]                       # ingresamos el centro, el vehiculo e iniciamos ruta
        veh_cap = np.copy(capacidad_vehiculos_copy[vehiculo_temp-1, :])   # copiamos la capacidad del vehiculo para evitar modificar la capacidad orginal
        while idx_c < len(asignados):                                     # mientras no se hayan recorrido todos los asignados(cliente o centro segun corresponda)
            if periodo > 1 and escalon == 1:
                dem_c = demanda[idx_c, 0:n_productos]                     # obtenemos la demanda del asignado
            else:
                dem_c = demanda[idx_c, rango:rango+n_productos]           # obtenemos la demanda del asignado
            resta = veh_cap - dem_c                                       # restamos la capacidad del vehiculo con la demanda del asignado
            binvec = np.array([binarize(x) for x in resta])               # binarizamos la resta
            if binvec.all():                                              # si la resta es positiva para cada producto
                ruta_temp.append(asignados[idx_c])                        # agregue a la ruta el centro accediendo al indice del mapeo que le corresponde
                veh_cap -= dem_c                                          # restamos la capacidad del vehiculo
                idx_c += 1                                                # aumentamos el indice
            else:                                                         # al agotar la capacidad del vehiculo
                if len(ruta_temp) != 0:
                    if ruta_temp[-1] == 0:                                    # si el vehiculo seleccionado no satisface ningun cliente
                        ruta_temp.pop(-1)                                     # elimina el 0 de inicio de ruta de ese vehiculo
                        ruta_temp.pop(-1)                                     # elimina el vehiculo de la ruta
                        vehiculos.pop(vehiculos.index(vehiculo_temp))         # eliminamos el vehiculo asigando de la lista de vehiculos
                        vehiculo_temp = np.random.choice(vehiculos)           # seleccionamos un nuevo vehiculo aleatorio
                        veh_cap = capacidad_vehiculos_copy[vehiculo_temp - 1, :]  # obtenemos la capacidad del nuevo vehiculo
                        ruta_temp += [vehiculo_temp, 0]                       # agregamos el vehiculo al plan de rutas e iniciamos una nueva ruta
                    else:
                        ruta_temp.append(0)                                   # finalizamos ruta
                        vehiculos.pop(vehiculos.index(vehiculo_temp))         # eliminamos el vehiculo asigando de la lista de vehiculos
                        vehiculo_temp = np.random.choice(vehiculos)           # seleccionamos un nuevo vehiculo aleatorio
                        veh_cap = capacidad_vehiculos_copy[vehiculo_temp - 1, :]   # obtenemos la capacidad del nuevo vehiculo
                        ruta_temp += [vehiculo_temp, 0]                       # agregamos el vehiculo al plan de rutas e iniciamos una nueva ruta
                else:
                    vehiculo_temp = np.random.choice(vehiculos)  # seleccionamos un vehiculo aleatorio
                    ruta_temp = [int(centro), vehiculo_temp, 0]  # ingresamos el centro, el vehiculo e iniciamos ruta
                    veh_cap = np.copy(capacidad_vehiculos_copy[vehiculo_temp - 1, :])

        vehiculos.pop(vehiculos.index(vehiculo_temp))                     # eliminamos el vehiculo de la lista de vehiculos
        ruta_temp.append(0)                                               # finalizamos ruta
        rutas_lv.append(ruta_temp)                                        # agregamos la ruta completa a la lista de rutas
    rutas_lv.append(-1)                                                   # agregamos un -1 que nos indica el salto de periodo

    return rutas_lv

# Funcion para la generacion de un individuo completo
# se recibe como parametros
# n_clientes : numero de clientes
# n_centros_locales: numero de centros locales
# n_centros_regionales: numero de centros regionales
# n_periodos: numero de periodos
# n_productos: nuymero de productos
# n_vehiculos_s: numero de vehiculos de segundo escalon
# n_vehiculos_p: numero de vehiculos de primer escalon
# capacidad_cl: matriz con las capacidades de los centros locales por producto
# capacidad_cr: matriz con las capacidades de los centros regionales por producto
# capacidad_vehiculos_p: matriz de capacidad de carga de los vehiculos de primer escalon por producto
# capacidad_vehiculos_s: matriz de capacidad de carga de los vehiculos de segundo escalon por producto
# demanda_cl: matriz de demanda de los clientes por producto y periodo
# La funcion devuelve o retorna:
# asignaciones_primer_lv: matriz de asignaciones de primer escalon
# asignaciones_segundo_lv: matriz de asignaciones de segundo escalon
# rutas_primer_lv: lista de listas con las rutas de primer escalon
# rutas_segundo_lv: lista de listas con las rutas de segundo escalon
# demanda_cr_full: lista de diccionarios con las demandas de cada centro regional por producto, cada elemento de la lista corresponde a un periodo


def individuo(n_clientes, n_centroslocales, n_centrosregionales, n_periodos, n_productos, n_vehiculos_s, n_vehiculos_p, capacidad_cl, capacidad_cr, capacidad_vehiculos_p, capacidad_vehiculos_s, demanda_clientes):

    asignaciones_segundo_lv = []
    rutas_segundo_lv = []
    asignaciones_primer_lv = []
    rutas_primer_lv = []
    cl_habs = []
    cr_habs = []
    demanda_cr_full = []
    demandas_cl = []

    for perioactual in range(1, n_periodos + 1):
        # llamada a la funcion de asignacion para la asignacion-localizacion del segundo escalon
        asignacion_segundo_nivel, demanda_cl = asignaciones(n_clientes, n_centroslocales, perioactual, n_productos, capacidad_cl, demanda_clientes, cl_habs, 2)
        # variables de mapeo de los centros locales habilitados en el segundo escalon
        n_cl_habs, demanda_cl_np, cl_habs = maping(demanda_cl)
        # llamada a la funcion de rutas para la creacion del plan de rutas del segundo escalon
        rutas_segundo_nivel = rutas(asignacion_segundo_nivel, n_vehiculos_s, capacidad_vehiculos_s, demanda_clientes, perioactual, n_productos, 2)
        # llamada a la funcion de asignacion para la asignacion-localizacion del primer escalon
        asignacion_primer_nivel, demanda_cr = asignaciones(n_cl_habs, n_centrosregionales, perioactual, n_productos, capacidad_cr, demanda_cl_np, cr_habs, 1)
        # reemplazamos los indices de los centros locales de la asignacion de primer escalon por los centros locales habilitados realmente
        asignacion_primer_nivel[0, :] = cl_habs
        # variables de mapeo de los centros locales habilitados en el primer escalon
        n_cr_habs, demanda_cr_np, cr_habs = maping(demanda_cr)
        # llamada a la funcion de rutas para la creacion del plan de rutas del primer escalon
        rutas_primer_nivel = rutas(asignacion_primer_nivel, n_vehiculos_p, capacidad_vehiculos_p, demanda_cl_np, perioactual, n_productos, 1)
        # adjuntamos las asignaciones y las rutas a las variables de retorno
        asignaciones_segundo_lv.append(asignacion_segundo_nivel)
        rutas_segundo_lv = rutas_segundo_lv + rutas_segundo_nivel
        asignaciones_primer_lv.append(asignacion_primer_nivel)
        rutas_primer_lv = rutas_primer_lv + rutas_primer_nivel
        # extraemos en un diccionario las demandas de los centros regionales por periodo
        demanda_cr_dict = dictionarize_cr(demanda_cr)
        demanda_cr_full.append(demanda_cr_dict)
        demandas_cl.append(demanda_cl)

    return asignaciones_primer_lv, asignaciones_segundo_lv, rutas_primer_lv, rutas_segundo_lv, demanda_cr_full, demandas_cl


# Funcion para la generacion de los valores de Q y de I que permiten llevar una gestion del inventario
# se reciben como parametros: el diccionario de las demandas de los centros regionales, el numero de periodos, el numero de productos, numero de centros regionales,
# la matriz de capacidad de los centros regionales, y los respectivos inventarios iniciales
# La funcion retorna:
# valoresQ: diccionario con los valores de Q calculados de cada centro regional para cada producto en cada periodo
# valoresI: diccionario con los valores de I calculados de cada centro regional para cada producto en cada periodo
def fun_inventario(demandas_cr_full, n_periodos, n_productos, n_centrosregionales, capacidad_cr, inventario):

    matriz_demanda = []                                                                             # inicializamos una matriz de demandas que contendra los valores de las demandas de cada cr en cada periodo por cada producto
    for p in range(n_periodos):
        dic_i = demandas_cr_full[p]                                                                 # demandas de los centros regionales en el periodo p
        matriz_demanda.append([])
        matriz_demanda[p] = [np.zeros([1, n_productos]) for _ in range(n_centrosregionales)]        # rellenamos de ceros la matriz de demanda en la posicion p del periodo actual con longitud de la cantidad de productos
        matriz_demanda[p] = [x[0] for x in matriz_demanda[p]]                                       # asignamos los valores de las demandas en el periodo p en la matriz de demandas
        for j, k in dic_i.items():
            matriz_demanda[p][int(j - 1)] = np.array(k)                                             # extraemos los keys y values del diccionario y los asignamos en la matriz de demandas
    matriz_demanda = np.array(matriz_demanda)                                                       # conversion de la matriz en numpy array
    centroshabs = []
    for c in range(n_centrosregionales):
        if np.sum(matriz_demanda[:, c, :]):
            centroshabs.append(c)                                                                   # almacenamos los centros regionales habilitados
    centroshabs = np.array(centroshabs)
    matriz_demanda = matriz_demanda[:, centroshabs, :]                                              # extraemos las demandas de la matriz de demandas de los centros habilitados
    valoresQ = {}
    valoresI = {}
    # para cada centro regional
    for u in range(len(centroshabs)):
        valoresQ[centroshabs[u] + 1] = []
        valoresI[centroshabs[u] + 1] = []
        # para cada producto
        for l in range(n_productos):
            pr_per = matriz_demanda[:, u, l]                                                        # lista de productos por periodo del centro
            carga0 = np.sum(pr_per)                                                                 # carga inicial del centro regional
            cap_centro = capacidad_cr[centroshabs[u], l]                                            # capacidad del centro regional
            # a partir de este punto se sigue la logica del algoritmo de gestion de inventarios propuesta en la solucion
            for w in range(n_periodos):
                d_cl = pr_per[w]
                if w == 0:
                    carga = carga0
                    iper1 = inventario[centroshabs[u], l]
                    if (iper1 >= d_cl) and (carga > cap_centro):
                        if cap_centro + d_cl <= carga:
                            Q = np.random.randint(0, cap_centro + d_cl - iper1)
                        else:
                            Q = np.random.randint(0, carga - iper1)
                    elif (iper1 < d_cl) and (carga > cap_centro):
                        if cap_centro + d_cl <= carga:
                            Q = np.random.randint(d_cl - iper1, cap_centro + d_cl - iper1)
                        else:
                            Q = np.random.randint(d_cl - iper1, carga - iper1)
                    elif (iper1 >= d_cl) and (carga <= cap_centro):
                        Q = np.random.randint(0, carga - iper1)
                    elif (iper1 < d_cl) and (carga <= cap_centro):
                        if d_cl - iper1 == carga - iper1:
                            Q = d_cl - iper1
                        else:
                            Q = np.random.randint(d_cl - iper1, carga - iper1)
                    valoresQ[centroshabs[u] + 1].append(Q)
                    It = iper1 + Q - d_cl
                    valoresI[centroshabs[u] + 1].append(It)
                    iper1 = It
                    carga = carga - d_cl - iper1
                elif 1 <= w < n_periodos - 1:
                    if (iper1 >= d_cl) and (carga > cap_centro):
                        if cap_centro + d_cl <= carga:
                            Q = np.random.randint(0, cap_centro + d_cl - iper1)
                        else:
                            Q = np.random.randint(0, carga)
                    elif (iper1 < d_cl) and (carga > cap_centro):
                        if cap_centro + d_cl - iper1 <= carga:
                            Q = np.random.randint(d_cl - iper1, cap_centro + d_cl - iper1)
                        else:
                            Q = np.random.randint(d_cl - iper1, carga)
                    elif (iper1 >= d_cl) and (carga <= cap_centro):
                        if carga == 0:
                            Q = carga
                        else:
                            Q = np.random.randint(0, carga)
                    elif (iper1 < d_cl) and (carga <= cap_centro):
                        if d_cl - iper1 == carga:
                            Q = carga
                        else:
                            Q = np.random.randint(d_cl - iper1, carga)
                    valoresQ[centroshabs[u] + 1].append(Q)
                    It = iper1 + Q - d_cl
                    valoresI[centroshabs[u] + 1].append(It)
                    iper1 = It
                    carga = carga - Q
                elif w == n_periodos - 1:
                    Q = carga
                    valoresQ[centroshabs[u] + 1].append(Q)
                    It = iper1 + Q - d_cl
                    valoresI[centroshabs[u] + 1].append(It)

    return valoresQ, valoresI


# Funcion para evaluacion de costos de rutas

def costo_ruteo(rutas_lv, n_periodos, costo_rutas, n_centros, costo_vehiculos):
    costrut2 = 0
    costvehi = 0
    np.warnings.filterwarnings('ignore', category=np.VisibleDeprecationWarning)
    rutas2copia = np.copy(rutas_lv)
    for _ in range(n_periodos):
        ruta_periodo = []
        vehi_habs = []
        for ruta in rutas2copia:
            if ruta == -1:
                break
            else:
                ruta_periodo.append(ruta)
        rutas2copia = np.delete(rutas2copia, np.array(range(len(ruta_periodo) + 1)))
        ruticas = []
        for miniruta in ruta_periodo:
            ruta_f = [list(g) for k, g in groupby(miniruta, lambda x: x != 0) if k]
            ruticas.append(ruta_f)
        for rutica in ruticas:
            rutica_cl = []
            r_cl = rutica[0][0]
            veh_r = rutica[0][1]
            vehi_habs.append(veh_r)
            for i in range(1, len(rutica)):
                if i % 2 != 0:
                    rutica_cl.append(rutica[i])
                else:
                    vehi_habs.append(rutica[i][0])
            for c_visit in rutica_cl:
                if len(c_visit) == 1:
                    costrut2 += costo_rutas[r_cl - 1, n_centros + int(c_visit[0]) - 1]
                    costrut2 += costo_rutas[n_centros + int(c_visit[0]) - 1, r_cl - 1]
                else:
                    for z in range(len(c_visit)):
                        if z == 0:
                            costrut2 += costo_rutas[r_cl - 1, n_centros + int(c_visit[z]) - 1]
                        elif z == len(c_visit) - 1:
                            costrut2 += costo_rutas[n_centros + int(c_visit[z - 1]) - 1, n_centros + int(c_visit[z]) - 1]
                            costrut2 += costo_rutas[n_centros + int(c_visit[z]) - 1, r_cl - 1]
                        else:
                            costrut2 += costo_rutas[n_centros + int(c_visit[z - 1]) - 1, n_centros + int(c_visit[z]) - 1]
        for vehic in vehi_habs:
            costvehi += costo_vehiculos[vehic-1]

    return costrut2, costvehi


# funcion de fitness para hallar los costos de F1
def fitness_f1(n_periodos, n_productos, asignaciones_segundo_lv, costo_instalaciones_cl, costo_instalaciones_cr, costo_compraproductos, costo_transporte, costo_inventario, valoresQ, valoresI,  rutas_segundo_lv, rutas_primer_lv, costo_rutas_s, costo_rutas_p, n_centroslocales, n_centrosregionales, costo_vehiculos_s, costo_vehiculos_p):
    cr_habs = list(valoresQ.keys())
    cl_habs = []
    for p in range(n_periodos):
        for c in asignaciones_segundo_lv[p][1]:
            if c not in cl_habs:
                cl_habs.append(int(c))
    cost_loc_cr = 0
    cost_loc_cl = 0
    costprod = 0
    costtrans = 0
    costinv = 0
    costrut2, costveh2 = costo_ruteo(rutas_segundo_lv, n_periodos, costo_rutas_s, n_centroslocales, costo_vehiculos_s)
    costrut1, costveh1 = costo_ruteo(rutas_primer_lv, n_periodos, costo_rutas_p, n_centrosregionales, costo_vehiculos_p)
    for cl in cl_habs:
        cost_loc_cl += costo_instalaciones_cl[cl - 1]
    for cr in cr_habs:
        cost_loc_cr += costo_instalaciones_cr[cr - 1]
        matrizQ = np.array(valoresQ[cr])
        matrizQp = np.array_split(matrizQ, n_periodos)
        matrizI = np.array(valoresI[cr])
        matrizIp = np.array_split(matrizI, n_periodos)
        for pr in range(n_productos):
            for t in range(n_periodos):
                costprod += matrizQp[t][pr] * costo_compraproductos[pr, t]
                costtrans += matrizQp[t][pr] * costo_transporte[cr - 1, (n_periodos * t) + pr]
                costinv += matrizIp[t][pr] * costo_inventario[cr - 1, pr]

    return cost_loc_cl, cost_loc_cr, costprod, costtrans, costinv, costrut2, costrut1, costveh2, costveh1


# funcion de fitness para hallar los costos de f2
def fitness_f2(rutas_lv, n_periodos, costo_humano, n_centros):
    costhum = 0
    np.warnings.filterwarnings('ignore', category=np.VisibleDeprecationWarning)
    rutascopia = np.copy(rutas_lv)
    for w in range(n_periodos):
        ruta_periodo = []
        costo_humanitario = costo_humano[w]
        for ruta in rutascopia:
            if ruta == -1:
                break
            else:
                ruta_periodo.append(ruta)
        rutascopia = np.delete(rutascopia, np.array(range(len(ruta_periodo) + 1)))
        ruticas = []
        for miniruta in ruta_periodo:
            rutica = [list(g) for k, g in groupby(miniruta, lambda x: x != 0) if k]
            ruticas.append(rutica)
        for rutica in ruticas:
            rutica_cl = []
            r_cl = rutica[0][0]
            for i in range(1, len(rutica)):
                if i % 2 != 0:
                    rutica_cl.append(rutica[i])
            for c_visit in rutica_cl:
                if len(c_visit) == 1:
                    costhum += costo_humanitario[r_cl - 1, n_centros + int(c_visit[0]) - 1]
                    costhum += costo_humanitario[n_centros + int(c_visit[0]) - 1, r_cl - 1]
                else:
                    for z in range(len(c_visit)):
                        if z == 0:
                            costhum += costo_humanitario[r_cl - 1, n_centros + int(c_visit[z]) - 1]
                        elif z == len(c_visit) - 1:
                            costhum += costo_humanitario[n_centros + int(c_visit[z - 1]) - 1, n_centros + int(c_visit[z]) - 1]
                            costhum += costo_humanitario[n_centros + int(c_visit[z]) - 1, r_cl - 1]
                        else:
                            costhum += costo_humanitario[n_centros + int(c_visit[z - 1]) - 1, n_centros + int(c_visit[z]) - 1]
    return costhum


# operador de seleccion
def selection(poblacion, fitness):
    ind_selec = []
    ind_rejec = []
    for _ in range(int(poblacion/2)):
        rand_1 = np.random.randint(poblacion)
        while rand_1 in ind_rejec or rand_1 in ind_selec:
            rand_1 = np.random.randint(poblacion)
        rand_2 = np.random.randint(poblacion)
        while (rand_2 in ind_rejec or rand_2 in ind_selec) and rand_2 != rand_1:
            rand_2 = np.random.randint(poblacion)
        if fitness[rand_1] < fitness[rand_2]:
            ind_selec.append(rand_1)
            ind_rejec.append(rand_2)
        else:
            ind_selec.append(rand_2)
            ind_rejec.append(rand_1)
    return ind_selec

# documentar inventarios, lineas adicionales y fitness
# realizar operadores geneticos

